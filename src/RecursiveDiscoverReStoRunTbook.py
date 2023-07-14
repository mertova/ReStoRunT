from openpyxl import load_workbook
import re
import argparse

parser = argparse.ArgumentParser(
    description='Discover in a recording which cells and sheets make up an ensemble of ReStoRunT sheets that can be reused.')
#parser.add_argument("--infile", action="store", required=True)
parser.add_argument("--inrecording", action="store", required=True)
parser.add_argument("--outbook", action="store", required=True)

args = parser.parse_args()

inrecording = vars(args)["inrecording"]
outbook = vars(args)["outbook"]

# load the workbook from the infile
inrecording = load_workbook(inrecording)

print("File read from %s" % (inrecording))

# Create directed graph: Which sheet links which other sheet?
# Find root



# find the restorunt sheet in the sheetfile
restorunt_sheet = None

to_be_removed = []
for potential_restorunt_sheet in inbook:
    if potential_restorunt_sheet.title.endswith(sheetname):
        print("this is the one! %s" % potential_restorunt_sheet.title)
    else:
        to_be_removed.append(potential_restorunt_sheet)

for sheet in to_be_removed:
    inbook.remove(sheet)

for potential_restorunt_sheet in inbook:
    if potential_restorunt_sheet.title == sheetname:
        for row in potential_restorunt_sheet.iter_rows():
            for cell in row:
                cell.value = 1234

# save the modified book to the outfile
inbook.save(outfilename)
print("File successfully written to %s" % outfilename)
