from openpyxl import load_workbook
import re
import argparse

parser = argparse.ArgumentParser(
    description='Add ReStoRunT sheet to existing workbook.')
#parser.add_argument("--infile", action="store", required=True)
parser.add_argument("--sheetfile", action="store", required=True)
parser.add_argument("--infile", action="store", required=True)
parser.add_argument("--resultfile", action="store", required=True)
parser.add_argument("--targetsheetname", action="store", required=True)

args = parser.parse_args()

infilename = vars(args)["infile"]
sheetfilename = vars(args)["sheetfile"]
outfilename = vars(args)["resultfile"]
targetsheetname = vars(args)["targetsheetname"]

# load the workbook from the infile
inbook = load_workbook(infilename)
sheetbook = load_workbook(sheetfilename)

print("File read from %s , %s" % (infilename, sheetbook))


def fix_coordinate(in_coordinate):
    """Turn an existing coordinate, for example A1 into $A$1, a in_coordinate
    that will not change when being copied elsewhere.
    """
    m = re.match(r'([A-Z]+)([0-9]+)', in_coordinate)

    if m:
        return "$%s$%s" % (m.groups(0)[0], m.groups(0)[1])
    else:
        return "Could not parse: %s" % in_coordinate


# find the restorunt sheet in the sheetfile
restorunt_sheet = None
original_name = ""
for potential_restorunt_sheet in sheetbook:
    m = re.match(r'^ReStoRunT-(.*)', potential_restorunt_sheet.title)
    if m:
        restorunt_sheet = potential_restorunt_sheet
        original_name = m.groups(1)[0]

# find the destination sheet
# rename the sheet references
# add the restorunt sheet into the inbook using
# https://stackoverflow.com/questions/42344041/how-to-copy-worksheet-from-one-workbook-to-another-one-using-openpyxl
# write the resulting book

# https://stackoverflow.com/questions/42344041/how-to-copy-worksheet-from-one-workbook-to-another-one-using-openpyxl
#  xl1 = openpyxl.load_workbook('workbook1.xlsx')
#  sheet you want to copy
#  s = openpyxl.load_workbook('workbook2.xlsx').active
#  s._parent = xl1
#  xl1._add_sheet(s)
#  xl1.save('some_path/name.xlsx')

for row in restorunt_sheet.iter_rows():
    for cell in row:
        v = cell.value
        if isinstance(v, str) and v.startswith("="):
            # formula
            v = v.replace(original_name, targetsheetname)
            print("Replaced: %s by %s -> %s" %
                  (original_name, targetsheetname, v))
            cell.value = v

        # copy all sheets
for sheet in inbook:
    if sheet.title == targetsheetname:
        print("found")
        restorunt_sheet._parent = inbook
        inbook._add_sheet(restorunt_sheet)
        for row in restorunt_sheet.iter_rows():
            for cell in row:
                print(cell.value)
        print("added %s" % restorunt_sheet.title)

# save the modified book to the outfile
inbook.save(outfilename)
print("File successfully written to %s" % outfilename)
