from openpyxl import load_workbook
import re
import argparse

parser = argparse.ArgumentParser(
    description='Add ReStoRunT sheet to existing workbook.')
parser.add_argument("--infile", action="store", required=True)
parser.add_argument("--outfile", action="store", required=True)

args = parser.parse_args()

infilename = vars(args)["infile"]

# load the workbook from the infile
inbook = load_workbook(infilename)


print("File read from %s" % infilename)

outfilename = vars(args)["outfile"]


def fix_coordinate(in_coordinate):
    """Turn an existing coordinate, for example A1 into $A$1, a in_coordinate
    that will not change when being copied elsewhere.
    """
    m = re.match(r'([A-Z]+)([0-9]+)', in_coordinate)

    if m:
        return "$%s$%s" % (m.groups(0)[0], m.groups(0)[1])
    else:
        return "Could not parse: %s" % in_coordinate


# copy all sheets
for sheet in inbook:
    if not re.match(r'^ReStoRunT', sheet.title):
        # if the sheet name does not start with ReStoRunT, then
        # we can be reasonably sure that it is not a ReStoRunT
        # sheet. As a consequence, we will create a ReStoRunT
        # sheet for the given sheet
        new_sheet = inbook.copy_worksheet(sheet)
        new_sheet.title = "ReStoRunT-%s" % sheet.title

        for row in new_sheet.iter_rows():
            for cell in row:
                new_sheet[cell.coordinate].value = "=%s!%s" % (
                    sheet.title, fix_coordinate(cell.coordinate))

# save the modified book to the outfile
inbook.save(outfilename)
print("File successfully written to %s" % outfilename)
