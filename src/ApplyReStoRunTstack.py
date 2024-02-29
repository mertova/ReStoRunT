import openpyxl 
import re
import argparse
import os

import openpyxl
from copy import copy

# https://stackoverflow.com/a/68800310
# CC-BY-SA 4.0

def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cell values and styles
    copy_sheet_attributes(source_sheet, target_sheet)

def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)


parser = argparse.ArgumentParser(
    description='Apply a folder of ReStoRunT workbooks (the stackdir!) to a folder of workbooks the sourcedir.')
#parser.add_argument("--infile", action="store", required=True)
parser.add_argument("--sourcedir", action="store", required=True)
parser.add_argument("--stackdir", action="store", required=True)
parser.add_argument("--resultfile", action="store", required=True)

args = parser.parse_args()

sourcedirname = vars(args)["sourcedir"]
stackdirname = vars(args)["stackdir"]
resultfilename = vars(args)["resultfile"]

print(f"""This program will read files from folder {sourcedirname}
in order to apply templates from {stackdirname} 
they will be stored in {resultfilename}
""")






#wb_target = openpyxl.Workbook()
#target_sheet = wb_target.create_sheet(..sheet_name..)

#wb_source = openpyxl.load_workbook(..path\\+\\file_name.., data_only=True)
#source_sheet = wb_source[..sheet_name..]

#copy_sheet(source_sheet, target_sheet)

#if 'Sheet' in wb_target.sheetnames:  # remove default sheet
#    
#wb_target.save('out.xlsx')

def get_books_in_filetree(startDirectory):
    result = list()
    for root, dirs, files in os.walk(startDirectory, topdown=False):
        for name in files:
            m = re.match(r'.*\.xlsx', name)
            if(m):
                result.append(os.path.join(root, name))
        # for name in dirs:
        #    print(os.path.join(root, name))
    return result;

stackfilenames = get_books_in_filetree(stackdirname)
sourcefilenames = get_books_in_filetree(sourcedirname)

stackbooks = [openpyxl.load_workbook(filename) for filename in stackfilenames ]

# restrict to one source file, so far.
# sourcefilenames = sourcefilenames[0:1]

destination_book = openpyxl.Workbook()

to_remove = set()

# maybe I need to do something different for international workbooks.
to_remove.add("Sheet")

original_names = dict()

stack_sheets = list()

for to_be_applied in stackbooks:
    for potential_restorunt_sheet in to_be_applied:
        m = re.match(r'^ReStoRunT-(.*)', potential_restorunt_sheet.title)
        if m:
            # this is a restorunt_sheet
            restorunt_sheet = potential_restorunt_sheet
            original_name = m.groups(1)[0]
    
            restorunt_sheet._parent = destination_book
            destination_book._add_sheet(restorunt_sheet)
            restorunt_sheet.title = "bp-%s" % restorunt_sheet.title   
            to_remove.add(restorunt_sheet.title)

            stack_sheets.append(restorunt_sheet)
            original_names[restorunt_sheet.title] = original_name


def apply_sheet(destination_book,source_book,restorunt_sheet,original_name):

    # if I want to apply the restorunt sheet multiple times,
    # I need to first copy it into the book

    print(restorunt_sheet.title)

    source_as_list = list(source_book)
    source_as_list.reverse()

    for sheet in source_as_list:
        # transfer sheet from initial book into result book
        # sheet._parent = destination_book
        # destination_book._add_sheet(sheet)
        target_sheet = destination_book.create_sheet(sheet.title)
        copy_sheet(sheet,target_sheet)
        print("Processing source sheet: %s" % sheet.title)

        if(not (sheet.title in original_names.keys())):

            # find out if this sheet applies
            applicable = True
            for row in restorunt_sheet.iter_rows():
                for cell in row:
                    if applicable:
                        m = re.match(r'(.*)::recognize',str(cell.value))
                        if(m):
                            coordinate = cell.coordinate
                            target_value = m.groups(1)[0]
                            # print("%s : %s matched %s"% (coordinate,cell.value,m.groups(1)[0]))
                            source_value = sheet[coordinate].value
                            
                            if(source_value!=target_value):
                                applicable = False

            if applicable: # check if the page from the stack is applicable
                print("Applicable: ")
                new_restorunt_sheet = destination_book.copy_worksheet(restorunt_sheet)
                #destination_book._add_sheet(new_restorunt_sheet)
                new_restorunt_sheet.title = "ReStoRunT-%s" % sheet.title
                print("Copied sheet as %s" % new_restorunt_sheet.title)
                print("Applicable: %s to %s" % (new_restorunt_sheet.title,sheet.title))
                for row in new_restorunt_sheet.iter_rows():
                    for cell in row:
                        v = cell.value
                        if isinstance(v, str) and v.startswith("="):
                            # formula
                            v = v.replace(original_name, sheet.title)
                            print("Replaced: %s by %s -> %s" %
                                (original_name, sheet.title, v))
                            cell.value = v
            else:
                print("This template %s does not apply to this sheet %s" % (restorunt_sheet.title,sheet.title))






def apply_stack(destination_book,source_book,stack_sheets,sourcefilenames):
    for restorunt_sheet in stack_sheets:
        original_name = original_names[restorunt_sheet.title]
        apply_sheet(destination_book,source_book,restorunt_sheet,original_name)

for source_filename in sourcefilenames: 
    source_book = openpyxl.load_workbook(source_filename)
    apply_stack(destination_book,source_book,stack_sheets,sourcefilenames)

for sheet_name in to_remove:
    print("Removing %s" % sheet_name)
    destination_book.remove(destination_book[sheet_name])
destination_book.save(resultfilename);
exit(0);
